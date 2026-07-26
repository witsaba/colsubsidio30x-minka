"""External / XLSX benchmark corpus loader (REQ-CORPUS-1..7).

Discovers per-contributor ``BD_AUDIOS.xlsx`` workbooks with sibling ``NOTAS_VOZ/``
audio folders, normalises numeric ``ID_UNICO`` without mutating source files, and
emits one ``CorpusClip`` per workbook row while failing the whole load on any
duplicate / missing / multi-match / unsupported-MIME / unlabeled-audio case.

The historical ``labels.csv`` layout is preserved verbatim; ``load_corpus`` is a
dispatcher that picks the XLSX adapter when at least one ``BD_AUDIOS.xlsx`` is
found and falls back to the CSV adapter otherwise.
"""

from __future__ import annotations

import csv
import json
import re
import unicodedata
from pathlib import Path
from typing import TypedDict

# Required schema for every workbook sheet. ``JSON PRODUCTOS`` is treated as
# opaque payload: present-but-unparsed. ``ACERTIVIDAD`` and ``DIFICULTAD`` are
# recognised but never allow-listed.
REQUIRED_XLSX_HEADERS = (
    "ID_UNICO",
    "TEXTO_AUDIO",
    "ACERTIVIDAD",
    "DIFICULTAD",
    "JSON PRODUCTOS",
)

# Extension → MIME. Closed table; non-audio extensions fail the loader.
MIME_BY_EXT: dict[str, str] = {
    ".ogg": "audio/ogg",
    ".webm": "audio/webm",
    ".wav": "audio/wav",
    ".mp3": "audio/mpeg",
}


class CorpusValidationError(ValueError):
    """Deterministic corpus validation failure."""


class CorpusClip(TypedDict, total=False):
    """One loaded clip.

    The set of keys depends on the loader: CSV keeps the historic contract
    (``clip_id``, ``condition``, ``transcript``, ``items``, ``is_garbage``)
    while XLSX adds ``dataset``, ``mime_type``, ``dificultad``,
    ``acertividad``. Every loader keeps ``audio_path``.
    """

    clip_id: str
    dataset: str
    audio_path: Path
    mime_type: str
    condition: str
    transcript: str
    items: list[str]
    is_garbage: bool
    dificultad: str
    acertividad: str


_TRUE_VALUES = {"true", "1", "yes", "y"}


# --- canonicalisation -------------------------------------------------------


def normalise_id(raw) -> str:
    """Render a workbook / filename id in its canonical 5-digit display form.

    Integral numeric values (``1.0``, ``"1"``, ``"00001"``) collapse to the same
    5-digit zero-padded key. Non-numeric ids are NFC-trimmed text. New numeric
    rows above 99999 keep their unpadded form so future growth stays collision-
    free; the historical convention only requires 5-digit padding for the
    current 1..N range.
    """

    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        value = int(raw)
    elif isinstance(raw, str):
        text = raw.strip()
        try:
            value = int(text)
        except ValueError:
            return unicodedata.normalize("NFC", text)
    else:
        text = str(raw).strip()
        try:
            value = int(text)
        except ValueError:
            return unicodedata.normalize("NFC", text)

    if 0 <= value <= 99999:
        return f"{value:05d}"
    return str(value)


_PADDED_NUMERIC = __import__("re").compile(r"^0{1,4}\d{1,5}$")


def audio_match_key(display_id: str) -> str:
    """Map a display id onto the audio file stem used in the filesystem.

    The current ``NOTAS_VOZ`` directories keep filenames unpadded (``1.ogg``,
    ``19.ogg``); the workbook ``ID_UNICO`` floats normalise to the zero-padded
    ``00001`` / ``00019`` display form. The match key strips the padding so the
    audio lookup can locate either form without source-file mutation.
    """

    if _PADDED_NUMERIC.fullmatch(display_id):
        return str(int(display_id))
    return display_id


# --- discovery --------------------------------------------------------------


_NOTAS_DIRNAME = "NOTAS_VOZ"


def discover_workbooks(root: Path) -> list[Path]:
    """Return every ``BD_AUDIOS.xlsx`` below ``root``, sorted relative path."""

    root = Path(root).resolve()
    if not root.exists():
        return []
    return sorted(p for p in root.rglob("BD_AUDIOS.xlsx") if p.is_file())


def _dataset_relpath(root: Path, dataset_dir: Path) -> str:
    rel = dataset_dir.relative_to(root)
    return rel.as_posix() if rel.parts else ""


def _resolve_audio(notas_dir: Path, display_id: str) -> tuple[Path, str]:
    """Find exactly one matching audio file; return ``(path, mime_type)``.

    Raises :class:`CorpusValidationError` for missing, multi-match, or
    unsupported extensions. The MIME table is closed; an unrelated extension
    is never silently coerced.
    """

    matches = sorted(
        path for path in notas_dir.glob(f"{display_id}.*") if path.is_file()
    )
    if not matches:
        raise CorpusValidationError(
            f"missing audio for id {display_id!r} in {notas_dir}"
        )
    if len(matches) > 1:
        names = ", ".join(p.name for p in matches)
        raise CorpusValidationError(
            f"multiple audio matches for id {display_id!r} in {notas_dir}: {names}"
        )

    audio_path = matches[0]
    mime = MIME_BY_EXT.get(audio_path.suffix.lower())
    if mime is None:
        raise CorpusValidationError(
            f"unsupported audio extension {audio_path.suffix!r} for id {display_id!r}"
        )
    return audio_path, mime


def _unreferenced_audio(
    workbook: Path, notas_dir: Path, claimed_ids: set[str]
) -> list[str]:
    """Return audio filenames that no workbook row references."""

    if not notas_dir.exists():
        return []
    claimed_stems = {claim for claim in claimed_ids}
    extras = []
    for path in sorted(notas_dir.iterdir()):
        if not path.is_file():
            continue
        if path.suffix.lower() not in MIME_BY_EXT:
            # Audio-shaped files only; unrelated extensions are validation-
            # separate and reported by the caller if relevant.
            continue
        if path.stem not in claimed_stems:
            extras.append(path.name)
    return extras


# --- XLSX adapter -----------------------------------------------------------


def load_xlsx_corpus(dataset_dir: Path) -> list[CorpusClip]:
    """Load one dataset directory (workbook + sibling ``NOTAS_VOZ/``)."""

    from openpyxl import load_workbook

    dataset_dir = Path(dataset_dir).resolve()
    workbook = dataset_dir / "BD_AUDIOS.xlsx"
    notas_dir = dataset_dir / _NOTAS_DIRNAME

    wb = load_workbook(workbook, read_only=True, data_only=True)
    try:
        sheet_name = wb.sheetnames[0]
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        raise CorpusValidationError(f"workbook {workbook} is empty")

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    missing = [col for col in REQUIRED_XLSX_HEADERS if col not in headers]
    if missing:
        raise CorpusValidationError(
            f"workbook {workbook.name} missing required columns: {', '.join(missing)}"
        )

    header_index = {name: index for index, name in enumerate(headers)}
    id_idx = header_index["ID_UNICO"]
    transcript_idx = header_index["TEXTO_AUDIO"]
    acertividad_idx = header_index["ACERTIVIDAD"]
    dificultad_idx = header_index["DIFICULTAD"]

    body = [tuple(row) for row in rows[1:] if any(cell not in (None, "") for cell in row)]

    # First pass: collect canonical ids + audio paths so a duplicate id or
    # missing audio fails before the unlabeled-audio check.
    clips: list[CorpusClip] = []
    seen_ids: dict[str, list[tuple[int, str]]] = {}
    claimed_stems: set[str] = set()
    for line_no, row in enumerate(body, start=2):
        raw_id = row[id_idx]
        if raw_id is None or str(raw_id).strip() == "":
            raise CorpusValidationError(
                f"workbook {workbook.name} row {line_no} has empty ID_UNICO"
            )
        display_id = normalise_id(raw_id)
        file_stem = audio_match_key(display_id)
        row_transcript = str(row[transcript_idx] or "").strip()
        acertividad = str(row[acertividad_idx] or "").strip()
        dificultad = str(row[dificultad_idx] or "").strip()

        if display_id not in seen_ids:
            seen_ids[display_id] = []
        seen_ids[display_id].append((line_no, row_transcript))
        claimed_stems.add(file_stem)

        try:
            audio_path, mime_type = _resolve_audio(notas_dir, file_stem)
        except CorpusValidationError as exc:
            raise CorpusValidationError(
                f"workbook {workbook.name} row {line_no} ({display_id}): {exc}"
            ) from exc

        clips.append(
            {
                "clip_id": _composite_id(dataset_dir, dataset_dir, display_id),
                "dataset": dataset_dir.name,
                "audio_path": audio_path,
                "mime_type": mime_type,
                "condition": "unknown",
                "transcript": row_transcript,
                "items": [],
                "is_garbage": False,
                "dificultad": dificultad,
                "acertividad": acertividad,
            }
        )

    duplicates = {
        display_id: occurrences for display_id, occurrences in seen_ids.items()
        if len(occurrences) > 1
    }
    if duplicates:
        rendered = {
            display_id: [f"row {ln} transcript={t!r}" for ln, t in occs]
            for display_id, occs in duplicates.items()
        }
        raise CorpusValidationError(
            f"workbook {workbook.name} has duplicate composite ids: {rendered}"
        )

    # Second pass: any audio in NOTAS_VOZ that no row claimed is unlabeled.
    extras = _unreferenced_audio(workbook, notas_dir, claimed_stems)
    if extras:
        raise CorpusValidationError(
            f"workbook {workbook.name}: unlabeled audio in NOTAS_VOZ: {', '.join(extras)}"
        )

    return clips


def _composite_id(root: Path, dataset_dir: Path, display_id: str) -> str:
    rel = _dataset_relpath(root, dataset_dir)
    return f"{rel}/{display_id}" if rel else display_id


# --- CSV adapter (labels.csv path preserved for legacy tests) --------------


def load_labels_csv(corpus_dir: Path) -> list[CorpusClip]:
    """Read ``labels.csv`` and pair every row with its sibling audio file.

    The shape mirrors the historic harness: ``clip_id``, ``condition``,
    ``transcript``, ``items``, ``is_garbage``, ``audio_path``. ``dataset`` is
    always the empty string for the CSV layout; ``mime_type`` is derived from
    the audio extension through the closed MIME table.
    """

    corpus_dir = Path(corpus_dir).resolve()
    labels = corpus_dir / "labels.csv"
    if not labels.exists():
        raise CorpusValidationError(f"no labels.csv in {corpus_dir}")

    with labels.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))

    clips: list[CorpusClip] = []
    for row in rows:
        clip_id = (row.get("clip_id") or "").strip()
        if not clip_id:
            continue
        audio_path, mime_type = _resolve_audio(
            corpus_dir, clip_id
        )
        clips.append(
            {
                "clip_id": clip_id,
                "dataset": "",
                "audio_path": audio_path,
                "mime_type": mime_type,
                "condition": (row.get("condition") or "").strip() or "unknown",
                "transcript": row.get("transcript") or "",
                "items": json.loads(row.get("items") or "[]"),
                "is_garbage": (row.get("is_garbage") or "").strip().lower() in _TRUE_VALUES,
            }
        )
    return clips


# --- dispatch ---------------------------------------------------------------


def load_corpus(root: Path) -> list[CorpusClip]:
    """Load the corpus under ``root``.

    Picks the XLSX adapter when ``root`` (or anything below it) contains a
    ``BD_AUDIOS.xlsx``; otherwise falls back to ``labels.csv``. A directory
    with neither is a validation error so a silent empty result never
    masquerades as a clean run.
    """

    root = Path(root).resolve()
    workbooks = discover_workbooks(root)
    if not workbooks:
        return load_labels_csv(root)

    all_clips: list[CorpusClip] = []
    for workbook in workbooks:
        dataset_dir = workbook.parent
        dataset_clips = load_xlsx_corpus(dataset_dir)
        rel_root = _dataset_relpath(root, dataset_dir)
        for clip in dataset_clips:
            clip["clip_id"] = (
                f"{rel_root}/{clip['clip_id']}" if rel_root else clip["clip_id"]
            )
            clip["dataset"] = rel_root or dataset_dir.name
        all_clips.extend(dataset_clips)
    return all_clips
