"""External / XLSX corpus loader (REQ-CORPUS-1..7, REQ-BMK-1).

This suite is the RED layer written before ``benchmarks/corpus_loader.py``. Every
test references symbols and behaviour that do not exist yet — the loader, the
``CorpusValidationError`` family, and the discovery helpers — so they fail to
collect until production code lands.
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

import pytest

# Unconditional import: the module does not exist yet, so RED collection fails
# at setup time. Once ``benchmarks/corpus_loader.py`` exists the suite asserts
# the behaviour the design promises.
from benchmarks.corpus_loader import (  # noqa: E402,F401  - RED import surface
    CorpusClip,
    CorpusValidationError,
    MIME_BY_EXT,
    discover_workbooks,
    load_corpus,
    load_labels_csv,
    load_xlsx_corpus,
    normalise_id,
)


# --- fixtures ---------------------------------------------------------------


@pytest.fixture
def write_xlsx():
    """Tiny openpyxl writer for fixture workbooks.

    The writer mirrors the current ``BD_AUDIOS.xlsx`` schema exactly:
    ``ID_UNICO`` (float), ``TEXTO_AUDIO``, ``ACERTIVIDAD``, ``DIFICULTAD``,
    ``JSON PRODUCTOS``. New rows and columns are intentionally disallowed.
    """

    from openpyxl import Workbook

    def _write(path: Path, rows: Iterable[dict], sheet_name: str = "Sheet1") -> None:
        wb = Workbook()
        sheet = wb.active
        sheet.title = sheet_name
        sheet.append(
            ["ID_UNICO", "TEXTO_AUDIO", "ACERTIVIDAD", "DIFICULTAD", "JSON PRODUCTOS"]
        )
        for row in rows:
            sheet.append(
                [
                    row["ID_UNICO"],
                    row["TEXTO_AUDIO"],
                    row.get("ACERTIVIDAD", ""),
                    row.get("DIFICULTAD", ""),
                    row.get("JSON PRODUCTOS", "{}"),
                ]
            )
        wb.save(path)

    return _write


@pytest.fixture
def two_dataset_corpus(tmp_path: Path, write_xlsx) -> Path:
    """Two 19-row workbooks, each with sibling NOTAS_VOZ 1.ogg..19.ogg."""

    rows = []
    for index in range(1, 20):
        rows.append(
            {
                "ID_UNICO": float(index),
                "TEXTO_AUDIO": f"transcript for {index}",
                "ACERTIVIDAD": "irrelevante",
                "DIFICULTAD": "FACIL",
                "JSON PRODUCTOS": "{}",
            }
        )
    braejan = tmp_path / "Braejan"
    daniel = tmp_path / "Daniel"
    braejan.mkdir()
    daniel.mkdir()
    write_xlsx(braejan / "BD_AUDIOS.xlsx", rows)
    write_xlsx(daniel / "BD_AUDIOS.xlsx", rows)
    for dataset in (braejan, daniel):
        notas = dataset / "NOTAS_VOZ"
        notas.mkdir()
        for index in range(1, 20):
            (notas / f"{index}.ogg").write_bytes(b"fake-ogg")
    return tmp_path


@pytest.fixture
def csv_corpus(tmp_path: Path) -> Path:
    labels = tmp_path / "labels.csv"
    labels.write_text(
        "clip_id,condition,transcript,items,is_garbage\n"
        'clean-01,clean,tres kilos,"[""3""]",false\n'
        'noisy-01,noisy,noventa,"[""90""]",false\n'
        'garbage-01,noisy,,[],true\n',
        encoding="utf-8",
    )
    for clip_id in ("clean-01", "noisy-01", "garbage-01"):
        (tmp_path / f"{clip_id}.webm").write_bytes(b"fake")
    return tmp_path


# --- REQ-CORPUS-2 corpus-size-agnostic discovery ----------------------------


def test_two_datasets_produce_38_clips_without_a_hardcoded_count(
    two_dataset_corpus: Path,
) -> None:
    clips = load_corpus(two_dataset_corpus)
    assert len(clips) == 38


def test_adding_two_rows_grows_the_corpus(two_dataset_corpus: Path, write_xlsx) -> None:
    # Append rows 20 and 21 (with audio) to Braejan and re-load.
    from openpyxl import load_workbook

    path = two_dataset_corpus / "Braejan" / "BD_AUDIOS.xlsx"
    wb = load_workbook(path)
    sheet = wb.active
    sheet.append([20.0, "new row twenty", "irrelevante", "FACIL", "{}"])
    sheet.append([21.0, "new row twenty one", "irrelevante", "FACIL", "{}"])
    wb.save(path)
    (two_dataset_corpus / "Braejan" / "NOTAS_VOZ" / "20.ogg").write_bytes(b"x")
    (two_dataset_corpus / "Braejan" / "NOTAS_VOZ" / "21.ogg").write_bytes(b"x")

    clips = load_corpus(two_dataset_corpus)
    assert len(clips) == 40


def test_discovery_walks_below_the_supplied_root(two_dataset_corpus: Path) -> None:
    paths = discover_workbooks(two_dataset_corpus)
    relpaths = sorted(p.relative_to(two_dataset_corpus).as_posix() for p in paths)
    assert relpaths == ["Braejan/BD_AUDIOS.xlsx", "Daniel/BD_AUDIOS.xlsx"]


# --- REQ-CORPUS-3 deterministic composite IDs -------------------------------


def test_canonical_numeric_id_is_zero_padded_to_five_digits() -> None:
    assert normalise_id(1.0) == "00001"
    assert normalise_id(19.0) == "00019"
    assert normalise_id("1") == "00001"
    assert normalise_id("00001") == "00001"


def test_non_numeric_ids_are_trimmed_nfc_text() -> None:
    assert normalise_id("  caf\u00e9  ") == "café"


def test_repeated_local_ids_across_datasets_get_distinct_composite_ids(
    two_dataset_corpus: Path,
) -> None:
    clips = load_corpus(two_dataset_corpus)
    clip_ids = [clip["clip_id"] for clip in clips]
    # both datasets should have their own 00001..00019; assert both appear
    assert "Braejan/00001" in clip_ids
    assert "Daniel/00001" in clip_ids
    assert "Braejan/00019" in clip_ids
    assert "Daniel/00019" in clip_ids
    assert len(set(clip_ids)) == 38


def test_duplicate_composite_id_fails_deterministically(
    tmp_path: Path, write_xlsx
) -> None:
    dataset = tmp_path / "Daniel"
    dataset.mkdir()
    notas = dataset / "NOTAS_VOZ"
    notas.mkdir()
    # Two rows both normalised to 00001 — that cannot be a valid mapping.
    write_xlsx(
        dataset / "BD_AUDIOS.xlsx",
        [
            {
                "ID_UNICO": 1.0,
                "TEXTO_AUDIO": "first",
                "ACERTIVIDAD": "irrelevante",
                "DIFICULTAD": "FACIL",
                "JSON PRODUCTOS": "{}",
            },
            {
                "ID_UNICO": 1.0,
                "TEXTO_AUDIO": "second",
                "ACERTIVIDAD": "irrelevante",
                "DIFICULTAD": "FACIL",
                "JSON PRODUCTOS": "{}",
            },
        ],
    )
    (notas / "1.ogg").write_bytes(b"x")

    with pytest.raises(CorpusValidationError) as excinfo:
        load_xlsx_corpus(dataset)
    assert "00001" in str(excinfo.value)


# --- REQ-CORPUS-4 strict one-row/one-audio mapping --------------------------


def test_missing_audio_fails_with_the_missing_id(tmp_path: Path, write_xlsx) -> None:
    dataset = tmp_path / "Daniel"
    dataset.mkdir()
    notas = dataset / "NOTAS_VOZ"
    notas.mkdir()
    write_xlsx(
        dataset / "BD_AUDIOS.xlsx",
        [
            {
                "ID_UNICO": 5.0,
                "TEXTO_AUDIO": "audio cinco",
                "ACERTIVIDAD": "irrelevante",
                "DIFICULTAD": "FACIL",
                "JSON PRODUCTOS": "{}",
            }
        ],
    )
    # No matching audio file; only an unrelated one.
    (notas / "9.ogg").write_bytes(b"x")

    with pytest.raises(CorpusValidationError) as excinfo:
        load_xlsx_corpus(dataset)
    assert "00005" in str(excinfo.value)


def test_multiple_extension_matches_fail(tmp_path: Path, write_xlsx) -> None:
    dataset = tmp_path / "Daniel"
    dataset.mkdir()
    notas = dataset / "NOTAS_VOZ"
    notas.mkdir()
    write_xlsx(
        dataset / "BD_AUDIOS.xlsx",
        [
            {
                "ID_UNICO": 1.0,
                "TEXTO_AUDIO": "uno",
                "ACERTIVIDAD": "irrelevante",
                "DIFICULTAD": "FACIL",
                "JSON PRODUCTOS": "{}",
            }
        ],
    )
    (notas / "1.ogg").write_bytes(b"x")
    (notas / "1.webm").write_bytes(b"x")

    with pytest.raises(CorpusValidationError):
        load_xlsx_corpus(dataset)


def test_unlabeled_audio_file_fails(tmp_path: Path, write_xlsx) -> None:
    dataset = tmp_path / "Daniel"
    dataset.mkdir()
    notas = dataset / "NOTAS_VOZ"
    notas.mkdir()
    write_xlsx(
        dataset / "BD_AUDIOS.xlsx",
        [
            {
                "ID_UNICO": 1.0,
                "TEXTO_AUDIO": "uno",
                "ACERTIVIDAD": "irrelevante",
                "DIFICULTAD": "FACIL",
                "JSON PRODUCTOS": "{}",
            }
        ],
    )
    (notas / "1.ogg").write_bytes(b"x")
    (notas / "99.ogg").write_bytes(b"x")  # unreferenced

    with pytest.raises(CorpusValidationError) as excinfo:
        load_xlsx_corpus(dataset)
    assert "99.ogg" in str(excinfo.value)


def test_unsupported_audio_extension_fails(tmp_path: Path, write_xlsx) -> None:
    dataset = tmp_path / "Daniel"
    dataset.mkdir()
    notas = dataset / "NOTAS_VOZ"
    notas.mkdir()
    write_xlsx(
        dataset / "BD_AUDIOS.xlsx",
        [
            {
                "ID_UNICO": 1.0,
                "TEXTO_AUDIO": "uno",
                "ACERTIVIDAD": "irrelevante",
                "DIFICULTAD": "FACIL",
                "JSON PRODUCTOS": "{}",
            }
        ],
    )
    (notas / "1.txt").write_text("not audio")

    with pytest.raises(CorpusValidationError) as excinfo:
        load_xlsx_corpus(dataset)
    assert ".txt" in str(excinfo.value)
    assert "unsupported" in str(excinfo.value)


# --- REQ-CORPUS-5 MIME table ------------------------------------------------


def test_mime_table_is_fixed_and_audio_only() -> None:
    assert MIME_BY_EXT == {
        ".ogg": "audio/ogg",
        ".webm": "audio/webm",
        ".wav": "audio/wav",
        ".mp3": "audio/mpeg",
    }


# --- REQ-CORPUS-6 opaque ACERTIVIDAD / DIFICULTAD ---------------------------


def test_unknown_acertividad_is_preserved_verbatim(tmp_path: Path, write_xlsx) -> None:
    dataset = tmp_path / "Daniel"
    dataset.mkdir()
    notas = dataset / "NOTAS_VOZ"
    notas.mkdir()
    write_xlsx(
        dataset / "BD_AUDIOS.xlsx",
        [
            {
                "ID_UNICO": 1.0,
                "TEXTO_AUDIO": "uno",
                "ACERTIVIDAD": "nuevo-tipo-2027",
                "DIFICULTAD": "FACIL",
                "JSON PRODUCTOS": "{}",
            }
        ],
    )
    (notas / "1.ogg").write_bytes(b"x")

    clips = load_xlsx_corpus(dataset)
    assert len(clips) == 1
    assert clips[0]["acertividad"] == "nuevo-tipo-2027"


# --- REQ-CORPUS-7 optional acoustic condition ------------------------------


def test_absent_condition_default_is_unknown(two_dataset_corpus: Path) -> None:
    clips = load_corpus(two_dataset_corpus)
    assert all(clip["condition"] == "unknown" for clip in clips)


def test_condition_is_never_inferred_from_dificultad(tmp_path: Path, write_xlsx) -> None:
    dataset = tmp_path / "Daniel"
    dataset.mkdir()
    notas = dataset / "NOTAS_VOZ"
    notas.mkdir()
    write_xlsx(
        dataset / "BD_AUDIOS.xlsx",
        [
            {
                "ID_UNICO": 1.0,
                "TEXTO_AUDIO": "uno",
                "ACERTIVIDAD": "irrelevante",
                "DIFICULTAD": "DIFICIL",
                "JSON PRODUCTOS": "{}",
            }
        ],
    )
    (notas / "1.ogg").write_bytes(b"x")

    clips = load_xlsx_corpus(dataset)
    assert clips[0]["condition"] == "unknown"
    assert clips[0]["dificultad"] == "DIFICIL"


# --- Required columns -------------------------------------------------------


def test_missing_required_columns_fail(tmp_path: Path, write_xlsx) -> None:
    from openpyxl import Workbook

    dataset = tmp_path / "Daniel"
    dataset.mkdir()
    notas = dataset / "NOTAS_VOZ"
    notas.mkdir()
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"
    # Forget the JSON PRODUCTOS header — the schema contract requires it.
    sheet.append(["ID_UNICO", "TEXTO_AUDIO", "ACERTIVIDAD", "DIFICULTAD"])
    sheet.append([1.0, "uno", "irrelevante", "FACIL"])
    wb.save(dataset / "BD_AUDIOS.xlsx")
    (notas / "1.ogg").write_bytes(b"x")

    with pytest.raises(CorpusValidationError):
        load_xlsx_corpus(dataset)


# --- CSV backward compatibility --------------------------------------------


def test_csv_corpus_loads_with_existing_contract(csv_corpus: Path) -> None:
    clips = load_labels_csv(csv_corpus)
    assert [c["clip_id"] for c in clips] == ["clean-01", "noisy-01", "garbage-01"]
    assert clips[0]["items"] == ["3"]
    assert clips[2]["is_garbage"] is True


def test_load_corpus_dispatches_to_csv_when_present(csv_corpus: Path) -> None:
    clips = load_corpus(csv_corpus)
    # CSV path preserves the plain clip_id; dataset is empty.
    assert clips[0]["dataset"] == ""
    assert clips[0]["clip_id"] == "clean-01"


def test_load_corpus_dispatches_to_xlsx_when_no_labels_present(
    two_dataset_corpus: Path,
) -> None:
    clips = load_corpus(two_dataset_corpus)
    assert {clip["dataset"] for clip in clips} == {"Braejan", "Daniel"}


def test_xlsx_clips_carry_mime_type(two_dataset_corpus: Path) -> None:
    clips = load_corpus(two_dataset_corpus)
    assert all(clip["mime_type"] == "audio/ogg" for clip in clips)


def test_xlsx_clips_carry_extended_metadata(two_dataset_corpus: Path) -> None:
    clips = load_corpus(two_dataset_corpus)
    sample = next(c for c in clips if c["clip_id"] == "Braejan/00001")
    assert sample["dataset"] == "Braejan"
    assert sample["audio_path"].name == "1.ogg"
    assert sample["dificultad"] == "FACIL"
    assert sample["acertividad"] == "irrelevante"
    assert sample["transcript"] == "transcript for 1"
    # items is empty: the XLSX schema does not label quantity tokens.
    assert sample["items"] == []
